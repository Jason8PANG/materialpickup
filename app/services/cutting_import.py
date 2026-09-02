# -*- coding: utf-8 -*-
"""
裁线参考表 kr_cutting_ref 导入服务。

复用 scripts/import_cutting_db.py 的列索引映射逻辑，
但改为接收 io.BytesIO（file-like），供 /api/cutting-ref/import 上传接口调用。

- 核心字段映射为命名列（裁线调用参考用）
- 整行原始数据保留在 raw_data JSON（防信息丢失）
- 覆盖式导入：TRUNCATE 旧数据后重导（默认 truncate=True）
"""
import json
import openpyxl

from app.models import get_db_connection

SHEET = 'BOM'
BATCH = 500

CREATE_TABLE = """
CREATE TABLE IF NOT EXISTS kr_cutting_ref (
  id              BIGINT        NOT NULL AUTO_INCREMENT,
  finished_part   VARCHAR(128)  DEFAULT NULL COMMENT '成品料号',
  wire_part       VARCHAR(128)  DEFAULT NULL COMMENT '线材（套管）料号',
  wire_awg        VARCHAR(32)   DEFAULT NULL COMMENT '线材AWG',
  color           VARCHAR(64)   DEFAULT NULL COMMENT '颜色',
  qty_per_group   DECIMAL(10,2) DEFAULT NULL COMMENT '单组剪切数量',
  cut_length_mm   DECIMAL(12,2) DEFAULT NULL COMMENT '剪切长度(mm)',
  length_tol      VARCHAR(32)   DEFAULT NULL COMMENT '长度公差',
  cut_device      VARCHAR(64)   DEFAULT NULL COMMENT '剪切设备',
  device_no       VARCHAR(64)   DEFAULT NULL COMMENT '设备序号',
  strip_len_a     DECIMAL(10,2) DEFAULT NULL COMMENT '去皮尺寸A端(mm)',
  strip_tol_a     VARCHAR(32)   DEFAULT NULL COMMENT '去皮尺寸公差A端',
  strip_len_b     DECIMAL(10,2) DEFAULT NULL COMMENT '去皮尺寸B端(mm)',
  strip_tol_b     VARCHAR(32)   DEFAULT NULL COMMENT '去皮尺寸公差B端',
  term_a          VARCHAR(128)  DEFAULT NULL COMMENT 'A端端子料号',
  term_b          VARCHAR(128)  DEFAULT NULL COMMENT 'B端端子料号',
  raw_data        JSON          DEFAULT NULL COMMENT '整行原始数据',
  created_at      DATETIME      NOT NULL DEFAULT CURRENT_TIMESTAMP,
  PRIMARY KEY (id),
  INDEX idx_finished_part (finished_part),
  INDEX idx_wire_part (wire_part),
  INDEX idx_finished_wire (finished_part, wire_part)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
COMMENT='成品对应物料裁线数据库（备料MP1导入，裁线时调用参考）'
"""

# 列索引映射：Excel 列下标 -> (目标字段名, 类型)。类型 'str' 或 'dec'。
COLUMN_MAP = [
    (0, 'finished_part', 'str'),
    (1, 'wire_part', 'str'),
    (2, 'wire_awg', 'str'),
    (3, 'color', 'str'),
    (4, 'qty_per_group', 'dec'),
    (5, 'cut_length_mm', 'dec'),
    (6, 'length_tol', 'str'),
    (7, 'cut_device', 'str'),
    (8, 'device_no', 'str'),
    (12, 'strip_len_a', 'dec'),
    (13, 'strip_tol_a', 'str'),
    (18, 'strip_len_b', 'dec'),
    (19, 'strip_tol_b', 'str'),
    (34, 'term_a', 'str'),
    (44, 'term_b', 'str'),
]

FIELD_ORDER = [fld for _, fld, _ in COLUMN_MAP]

INSERT_SQL = (
    "INSERT INTO kr_cutting_ref "
    "(finished_part, wire_part, wire_awg, color, qty_per_group, cut_length_mm, "
    " length_tol, cut_device, device_no, strip_len_a, strip_tol_a, "
    " strip_len_b, strip_tol_b, term_a, term_b, raw_data) "
    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
)


def to_dec(v):
    if v is None or str(v).strip() == '':
        return None
    try:
        return float(str(v).strip().replace(',', ''))
    except (TypeError, ValueError):
        return None


def to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _convert(v, typ):
    return to_dec(v) if typ == 'dec' else to_str(v)


def import_cutting_ref(file_stream, sheet=SHEET, truncate=True):
    """从 file-like（io.BytesIO）读取 Excel BOM 表，覆盖式导入 kr_cutting_ref。

    参数:
        file_stream: io.BytesIO 或其它二进制可读对象（已定位到文件头）
        sheet:       工作表名，默认 'BOM'；不存在时回退到第一个工作表
        truncate:    True 时先 TRUNCATE 旧数据（覆盖式）

    返回:
        {'imported': int, 'skipped': int}
    """
    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).replace('\n', ' ').strip() if h else '' for h in next(rows_iter, [])]

    total = 0
    skipped = 0
    with get_db_connection() as db:
        cur = db.cursor()
        cur.execute(CREATE_TABLE)
        db.commit()
        if truncate:
            cur.execute('TRUNCATE TABLE kr_cutting_ref')
            db.commit()

        batch = []
        for row in rows_iter:
            vals = list(row)
            fp = to_str(vals[0]) if len(vals) > 0 else None
            wp = to_str(vals[1]) if len(vals) > 1 else None
            if not fp and not wp:
                skipped += 1
                continue  # 跳过完全空行
            raw = {header[i]: (str(v) if v is not None else None)
                   for i, v in enumerate(vals)
                   if i < len(header) and header[i] and v is not None}
            record = [_convert(vals[idx], typ) if len(vals) > idx else None
                      for idx, _, typ in COLUMN_MAP]
            record.append(json.dumps(raw, ensure_ascii=False) if raw else None)
            batch.append(tuple(record))
            total += 1
            if len(batch) >= BATCH:
                cur.executemany(INSERT_SQL, batch)
                db.commit()
                batch = []

        if batch:
            cur.executemany(INSERT_SQL, batch)
            db.commit()

    wb.close()
    return {'imported': total, 'skipped': skipped}
